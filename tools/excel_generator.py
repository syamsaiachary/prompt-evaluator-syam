# ─────────────────────────────────────────────
#  tools/excel_generator.py
#  Tool: generate_excel
#  Merges evaluation results with original CSV data
#  and writes a colour-coded annotated Excel file.
# ─────────────────────────────────────────────
from __future__ import annotations
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from langchain_core.tools import tool
from config import GRADE_COLOURS, OUTPUT_PATH


def _grade(total: int | None) -> str:
    if total is None: return "Flagged"
    if total >= 40: return "Excellent"
    if total >= 35: return "Good"
    return "Needs Improvement"


@tool
def  generate_excel(results: list[dict], original_rows: list[dict]) -> str:
    """
    Merges evaluation results with original CSV rows and writes
    a colour-coded annotated Excel file.

    Args:
        results:       List of scored dicts (one per submitted prompt).
        original_rows: Original CSV rows in the same order.

    Returns:
        Path to the generated Excel file.
    """
    os.makedirs(os.path.dirname(OUTPUT_PATH) or ".", exist_ok=True)

    # ── Build output dataframe ─────────────────
    records = []
    sorted_results = sorted(results, key=lambda r: r["index"])

    for item in sorted_results:
        s   = item["scores"]
        row = original_rows[item["index"]].copy()
        total_val = s.get("total")
        grade_str = _grade(total_val)
        display_score = f"{total_val}/50" if total_val is not None else "N/A"
        row.update({
            "Rubrics Split up": (
            f"Task: {s.get('task', {}).get('score', 0)}/10\n"
            f"Context: {s.get('context', {}).get('score', 0)}/10\n"
            f"Persona: {s.get('persona', {}).get('score', 0)}/10\n"
            f"Output: {s.get('output', {}).get('score', 0)}/10\n"
            f"Examples: {s.get('examples', {}).get('score', 0)}/4\n"
            f"About You: {s.get('about_you', {}).get('score', 0)}/4\n"
            f"Target Audience: {s.get('tg', {}).get('score', 0)}/3"),
            "Grade & Total Score": f"{grade_str} ({display_score})",
            "Feedback": f"{s.get('three_sentence_feedback','')}"
        })
        records.append(row)

    df = pd.DataFrame(records)
    df.to_excel(OUTPUT_PATH, index=False)

    # ── Apply colour coding via openpyxl ───────
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active

    # Header styling
    headers = [cell.value for cell in ws[1]]

    # Identify score-related columns (to receive color coding)
# Identify score-related columns (to receive color coding)
    score_col_indices = {
        i + 1
        for i, h in enumerate(headers)
        if h and str(h) == "Grade & Total Score"
    }
    grade_col_idx = headers.index("Grade & Total Score") + 1 if "Grade & Total Score" in headers else None

    # Border styles
    thin = Side(style="thin", color="D0D0D0")
    thick_bottom = Side(style="medium", color="1F1F1F")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=thin, right=thin, top=thin, bottom=thick_bottom)

    # ── Header row ────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3864")   # deep navy
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # ── Data rows ─────────────────────────────────────────
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Determine grade color (for score columns only)
        grade_raw = row[grade_col_idx - 1].value if grade_col_idx else "Flagged"
        grade_val = str(grade_raw or "Flagged").split(" (")[0]
        hex_colour = GRADE_COLOURS.get(grade_val, "FFFFFF")
        score_fill = PatternFill("solid", fgColor=hex_colour)

        # Zebra stripe for non-score columns
        zebra_fill = PatternFill("solid", fgColor="E0E0E0" if row_idx % 2 == 0 else "FFFFFF")

        for cell in row:
            col_idx = cell.column
            is_score_col = col_idx in score_col_indices

            cell.fill = score_fill if is_score_col else zebra_fill
            cell.border = thin_border
            cell.font = Font(
                name="Arial",
                size=9,
                bold=(is_score_col and str(cell.value or "").isdigit() is False and cell.column == grade_col_idx),
                color="1F1F1F",
            )
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if is_score_col else "left",
            )

        ws.row_dimensions[row_idx].height = 60

    # ── Column widths ─────────────────────────────────────
    for col_idx, col_cells in enumerate(ws.columns, 1):
        header_val = str(headers[col_idx - 1] or "")
        if col_idx in score_col_indices:
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        else:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH
 